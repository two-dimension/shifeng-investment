#!/usr/bin/env python3
"""从 input.json 生成业绩报告 Excel 和 PDF。"""
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


HERE = Path(__file__).resolve().parent
DETAIL_COLUMNS = [
    "证券代码", "证券简称", "所属子集", "报告类型", "报告期",
    "营业收入亿元", "营业收入同比%", "归母净利润亿元", "归母净利润同比%",
    "扣非净利润亿元", "扣非净利润同比%", "经营现金流亿元", "经营现金流同比%",
    "基本每股收益元", "加权ROE%", "指标覆盖数", "全文解析状态",
    "PDF总页数", "PDF已读页数", "公告标题", "原文链接", "公告ID",
]


def _number(value, digits: int = 2) -> str:
    if value is None or value == "":
        return "—"
    try:
        return f"{float(value):,.{digits}f}"
    except (TypeError, ValueError):
        return str(value)


def _pct(value) -> str:
    return "—" if value is None or value == "" else f"{float(value):,.2f}%"


def build_excel(data: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    items = data.get("items") or []
    summary = data.get("fetch_summary") or {}
    workbook = Workbook()
    overview = workbook.active
    overview.title = "概览"
    detail = workbook.create_sheet("明细")
    audit = workbook.create_sheet("抓取校验")
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)

    overview.append([f"A股业绩报告日报 · {data.get('date', '')}"])
    overview.merge_cells("A1:F1")
    overview["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    overview["A1"].fill = header_fill
    overview["A1"].alignment = Alignment(horizontal="center")
    overview_rows = [
        ("生成时间", data.get("generated_at", "")),
        ("数据来源", data.get("data_source", "")),
        ("正式报告总数", len(items)),
        ("巨潮去重公告数", summary.get("dedup_rows", 0)),
        ("全市场正式报告数", summary.get("formal_report_rows", 0)),
        ("其中 watchlist 报告数", summary.get("watchlist_report_rows", 0)),
        ("财务指标抽取成功", summary.get("metric_parse_ok", 0)),
    ]
    for label, value in overview_rows:
        overview.append([label, value])
    overview.column_dimensions["A"].width = 28
    overview.column_dimensions["B"].width = 48

    detail.append(DETAIL_COLUMNS)
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for item in items:
        detail.append([item.get(column) for column in DETAIL_COLUMNS])
    detail.freeze_panes = "C2"
    widths = [11, 13, 22, 15, 11, 14, 13, 16, 14, 16, 14, 16, 14, 14, 12, 11, 15, 11, 11, 34, 42, 17]
    for index, width in enumerate(widths, start=1):
        detail.column_dimensions[get_column_letter(index)].width = width
    for row in detail.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    audit.append(["检查项", "实际值", "预期值", "状态", "说明"])
    for cell in audit[1]:
        cell.fill = header_fill
        cell.font = header_font
    checks = [
        ("巨潮查询完整", bool(data.get("fetch_complete")), True, "任一巨潮分页失败即 FAIL"),
        ("明细数与纳入报告数一致", len(items), summary.get("included_report_rows", len(items)), "全部正式报告均进入明细"),
        ("指标成功数不超过明细数", summary.get("metric_parse_ok", 0), len(items), "抽取成功数必须落在有效区间"),
        ("抓取错误数", len(summary.get("errors") or []), 0, "抓取错误必须为 0"),
    ]
    for label, actual, expected, note in checks:
        ok = actual <= expected if label == "指标成功数不超过明细数" else actual == expected
        audit.append([label, actual, expected, "OK" if ok else "FAIL", note])
    audit.column_dimensions["A"].width = 32
    audit.column_dimensions["E"].width = 48
    workbook.save(output_path)


def _pdf_styles():
    candidates = [
        os.environ.get("RESEARCH_CJK_FONT"),
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    font_path = next((Path(value) for value in candidates if value and Path(value).is_file()), None)
    if font_path is None:
        raise FileNotFoundError(
            "未找到中文 PDF 字体；请安装 fonts-noto-cjk 或设置 RESEARCH_CJK_FONT"
        )
    if "HeitiSC" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("HeitiSC", str(font_path)))
        pdfmetrics.registerFont(TTFont("HeitiSCBold", str(font_path)))
        registerFontFamily(
            "HeitiSC",
            normal="HeitiSC",
            bold="HeitiSCBold",
            italic="HeitiSC",
            boldItalic="HeitiSCBold",
        )
    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "CN",
        parent=styles["BodyText"],
        fontName="HeitiSC",
        fontSize=8.5,
        leading=12,
        textColor=colors.HexColor("#243447"),
        alignment=TA_LEFT,
    )
    title = ParagraphStyle(
        "CNTitle",
        parent=base,
        fontSize=20,
        leading=26,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17365D"),
    )
    heading = ParagraphStyle(
        "CNHeading",
        parent=base,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#17365D"),
        spaceBefore=6,
        spaceAfter=6,
    )
    small = ParagraphStyle("CNSmall", parent=base, fontSize=7, leading=9)
    header_small = ParagraphStyle("CNHeaderSmall", parent=small, textColor=colors.white, alignment=TA_CENTER)
    return base, title, heading, small, header_small


def _draw_pdf_footer(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("HeitiSC", 7)
    canvas.setFillColor(colors.HexColor("#6B7280"))
    canvas.drawString(12 * mm, 6 * mm, "数据来源：巨潮资讯官方公告 · 财务指标为 PDF 自动抽取")
    canvas.drawRightString(landscape(A4)[0] - 12 * mm, 6 * mm, f"第 {doc.page} 页")
    canvas.restoreState()


def build_pdf(data: dict, output_path: Path) -> None:
    base, title_style, heading, small, header_small = _pdf_styles()
    items = data.get("items") or []
    summary = data.get("fetch_summary") or {}
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"A股业绩报告-{data.get('date', '')}",
        author="业绩报告自动任务",
    )
    story = [
        Paragraph(f"A股业绩报告日报 · {escape(str(data.get('date', '')))}", title_style),
        Spacer(1, 3 * mm),
        Paragraph(
            f"纳入报告 <b>{len(items)}</b> 份 · 全市场正式报告 {summary.get('formal_report_rows', 0)} 份 · "
            f"其中 watchlist {summary.get('watchlist_report_rows', 0)} 份 · "
            f"财务指标成功抽取 {summary.get('metric_parse_ok', 0)} 份",
            base,
        ),
        Spacer(1, 2 * mm),
    ]
    if not items:
        story.extend([
            Paragraph("今日无正式年度、半年度、一季或三季报告披露。", heading),
            Paragraph("空结果仍保留巨潮查询明细和校验信息，不能据此推断全市场无其他类型公告。", base),
        ])
    else:
        story.append(Paragraph("报告明细", heading))
        header = ["公司", "类型/期间", "营收(亿元)", "营收同比", "归母净利(亿元)", "归母同比", "扣非净利(亿元)", "扣非同比", "现金流(亿元)", "原文"]
        rows = [[Paragraph(f"<b>{escape(x)}</b>", header_small) for x in header]]
        for item in items:
            url = escape(str(item.get("原文链接") or ""), {'"': '&quot;'})
            company = f"{item.get('证券简称', '')}<br/>{item.get('证券代码', '')}"
            rows.append([
                Paragraph(escape(company).replace("&lt;br/&gt;", "<br/>"), small),
                Paragraph(f"{escape(str(item.get('报告类型', '')))}<br/>{escape(str(item.get('报告期', '')))}", small),
                Paragraph(_number(item.get("营业收入亿元")), small),
                Paragraph(_pct(item.get("营业收入同比%")), small),
                Paragraph(_number(item.get("归母净利润亿元")), small),
                Paragraph(_pct(item.get("归母净利润同比%")), small),
                Paragraph(_number(item.get("扣非净利润亿元")), small),
                Paragraph(_pct(item.get("扣非净利润同比%")), small),
                Paragraph(_number(item.get("经营现金流亿元")), small),
                Paragraph(f'<link href="{url}" color="#0563C1">原文</link>' if url else "—", small),
            ])
        table = Table(rows, colWidths=[28*mm, 29*mm, 25*mm, 22*mm, 28*mm, 22*mm, 28*mm, 22*mm, 27*mm, 15*mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E1F2")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)

    story.extend([Spacer(1, 4 * mm), Paragraph("口径与校验", heading)])
    for note in data.get("notes") or []:
        story.append(Paragraph(f"• {escape(str(note))}", base))
    story.append(Paragraph(
        f"生成时间：{escape(str(data.get('generated_at', datetime.now().isoformat())))}；"
        f"巨潮查询完整性：{'通过' if data.get('fetch_complete') else '失败'}。",
        small,
    ))
    doc.build(story, onFirstPage=_draw_pdf_footer, onLaterPages=_draw_pdf_footer)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_json", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()
    data = json.loads(args.input_json.read_text(encoding="utf-8"))
    date_str = data["date"]
    xlsx = args.output_dir / f"A股业绩报告-{date_str}.xlsx"
    pdf = args.output_dir / f"A股业绩报告-{date_str}.pdf"
    build_excel(data, xlsx)
    build_pdf(data, pdf)
    print(f"[build] {xlsx}")
    print(f"[build] {pdf}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
