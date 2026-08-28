#!/usr/bin/env python3
"""
processed_<date>.json → Excel 多 sheet 输出
- Brief:     概览(总数/利好/利空/集中行业/情绪总结)
- 每日全部利好:  所有命中利好的公告 (按分数排序)
- 每日全部利空:  所有命中利空的公告 (按分数排序)
- 多空全表:  全部利好/利空公司(各前 20)
- 排除清单:  中性过滤的公告原因汇总
- 信号规则:  Skill 14 条信号定义(便于查证)
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ===== 样式 =====
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Microsoft YaHei', size=14, bold=True, color='1F4E78')
H2_FONT = Font(name='Microsoft YaHei', size=12, bold=True, color='2E75B6')
BOLD = Font(name='Microsoft YaHei', size=10, bold=True)
NEUTRAL_FONT = Font(name='Microsoft YaHei', size=10)
# v9.7 配色反转: 利好红/利空绿 (A 股红涨绿跌 + 跟 PDF RED/GREEN 一致)
GOOD_FILL = PatternFill('solid', fgColor='FFE6E6')  # 利好: 浅红底
GOOD_FONT = Font(name='Microsoft YaHei', size=10, bold=True, color='C00000')  # 利好: 红字
BAD_FILL = PatternFill('solid', fgColor='E8F5E9')  # 利空: 浅绿底
BAD_FONT = Font(name='Microsoft YaHei', size=10, bold=True, color='2E7D32')  # 利空: 绿字
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
WARN_FONT = Font(name='Microsoft YaHei', size=10, color='9C5700')

# v9.2: 涨跌方向颜色 (A 股习惯: 涨红跌绿, 跟 PDF RED/GREEN 一致)
PCT_UP_FONT = Font(name='Microsoft YaHei', size=10, bold=True, color='C00000')
PCT_UP_FILL = PatternFill('solid', fgColor='FFE6E6')
PCT_DOWN_FONT = Font(name='Microsoft YaHei', size=10, bold=True, color='2E7D32')
PCT_DOWN_FILL = PatternFill('solid', fgColor='E8F5E9')
THIN = Border(
    left=Side(style='thin', color='B7B7B7'),
    right=Side(style='thin', color='B7B7B7'),
    top=Side(style='thin', color='B7B7B7'),
    bottom=Side(style='thin', color='B7B7B7'),
)


def style_header(ws, r: int, n: int) -> None:
    for c in range(1, n + 1):
        cell = ws.cell(r, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN


def style_data(ws, r: int, n: int, fill=None, font=None, align: str = 'left', skip_cols=None) -> None:
    for c in range(1, n + 1):
        if c in (skip_cols or []):
            continue
        cell = ws.cell(r, c)
        cell.border = THIN
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal=align)
        if fill:
            cell.fill = fill
        cell.font = font or NEUTRAL_FONT


def render(data: Dict[str, Any], output_path: Path, raw: List[Dict] = None) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    sentiment = data.get('sentiment', {})
    excluded = data.get('excluded_summary', [])
    neutral_ann = data.get('neutral_announcements', [])  # 中性公告明细
    top_good = data.get('top_good', [])
    top_bad = data.get('top_bad', [])
    all_good = data.get('all_good_companies', [])
    all_bad = data.get('all_bad_companies', [])
    short_term = data.get('short_term', {})
    score_dist = data.get('score_distribution', {})
    fetch_meta = data.get('fetch_meta', {})
    date_label = data.get('date', '?')

    # ==================== Sheet: Brief ====================
    ws = wb.create_sheet('Brief')
    ws.column_dimensions['A'].width = 4
    for col in 'BCDEF':
        ws.column_dimensions[col].width = 22
    ws.column_dimensions['D'].width = 35

    coverage = data.get('coverage', {})
    range_label = coverage.get('range_label', date_label)
    day_count = coverage.get('day_count', 1)
    per_day = data.get('per_day', {})
    actual_dates = coverage.get('actual_dates', [])  # v9.37: 实际有数据的日期 (raw announcementTime 分日)
    weekend_note = coverage.get('weekend_empty_note', '')  # v9.37: 周末 0 公告提示

    r = 1
    title_main = f"巨潮资讯 公告研判 (覆盖区间: {range_label}, {day_count} 天"
    if actual_dates and len(actual_dates) < day_count:
        title_main += f" · 实际数据日: {', '.join(actual_dates)}"
    title_main += ")"
    ws.cell(r, 1, title_main).font = Font(
        name='Microsoft YaHei', size=18, bold=True, color='1F4E78')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, "数据源: cninfo.com.cn | 生成时间: " + datetime.now().strftime('%Y-%m-%d %H:%M:%S')).font = Font(
        name='Microsoft YaHei', size=10, italic=True, color='808080')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    if weekend_note:
        # v9.37: 周末 0 公告黄字警示, 让 trader 一眼明白不是漏扫
        ws.cell(r, 1, weekend_note).font = Font(
            name='Microsoft YaHei', size=10, italic=True, color='C77700')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    r += 1

    # 覆盖区间详情 (v9.37: 列出扫描范围内全部日期, 0 条日期显式标 "(cninfo 0 公告, 周末常态)")
    if day_count >= 1:
        ws.cell(r, 1, "零、覆盖区间按日分布").font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
        ws.cell(r, 2, "扫描日期").font = HEADER_FONT
        ws.cell(r, 3, "覆盖条数").font = HEADER_FONT
        style_header(ws, r, 3)
        r += 1
        # 列出扫描范围内全部日期 (start_date ~ end_date), 0 条的显式标空
        from datetime import datetime as _dt, timedelta as _td
        _s = _dt.strptime(coverage.get('start_date', date_label), '%Y-%m-%d')
        all_dates = [(_s + _td(days=i)).strftime('%Y-%m-%d') for i in range(day_count)]
        for d in all_dates:
            cnt = per_day.get(d, 0)
            ws.cell(r, 2, d).font = BOLD
            ws.cell(r, 2).alignment = Alignment(horizontal='center')
            if cnt > 0:
                ws.cell(r, 3, cnt).font = NEUTRAL_FONT
            else:
                # v9.37: 0 条日期显式标 "(周末/节假日 cninfo 0 公告, 非漏扫)"
                _wd = _dt.strptime(d, '%Y-%m-%d').weekday()
                _wk = '周一' if _wd == 0 else '周二' if _wd == 1 else '周三' if _wd == 2 else '周四' if _wd == 3 else '周五' if _wd == 4 else '周六' if _wd == 5 else '周日'
                ws.cell(r, 3, f'0 ({_wk} cninfo 无公告)').font = Font(
                    name='Microsoft YaHei', size=10, italic=True, color='808080')
            ws.cell(r, 3).alignment = Alignment(horizontal='center')
            style_data(ws, r, 3)
            r += 1
        r += 1

    ws.cell(r, 1, "一、关键指标").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    metrics = [
        ('数据日期', date_label),
        ('公告总数', fetch_meta.get('total', 0)),
        ('利好条数', sentiment.get('good_count', 0)),
        ('利空条数(全口径)', sentiment.get('bad_count', 0)),
        ('  其中: 强利空(<=-7)', sentiment.get('strong_bad_count', 0)),
        ('中性(过滤)', sentiment.get('neutral_count', 0)),
        ('利好公司数(去重)', len(all_good)),
        ('利空公司数(去重, -7~-10)', len(all_bad)),
    ]
    for label, val in metrics:
        ws.cell(r, 2, label).font = H2_FONT
        cell = ws.cell(r, 3, val)
        cell.font = Font(name='Microsoft YaHei', size=11, bold=True, color='1F4E78')
        cell.alignment = Alignment(horizontal='center')
        r += 1

    r += 1
    ws.cell(r, 1, "二、市场情绪").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    if sentiment.get('summary'):
        ws.cell(r, 2, "全市场多空对比").font = H2_FONT
        ws.cell(r, 3, sentiment['summary']).font = NEUTRAL_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        r += 1
    # v9.30 (老板铁律): 删"利好集中赛道"/"利空高发板块" → 改"重大利好标的"/"重大利空标的"
    if data.get("top_good"):
        top_g = data["top_good"][:5]
        targets = " · ".join([f"{c['company']}({c['code']}) {c.get('best_score', 0):+d}" for c in top_g])
        ws.cell(r, 2, "重大利好标的").font = H2_FONT
        c = ws.cell(r, 3, targets)
        c.font = GOOD_FONT
        c.fill = GOOD_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 36
        r += 1
    if data.get("top_bad"):
        top_b = data["top_bad"][:5]
        targets = " · ".join([f"{c['company']}({c['code']}) {c.get('best_score', 0):+d}" for c in top_b])
        ws.cell(r, 2, "重大利空标的").font = H2_FONT
        c = ws.cell(r, 3, targets)
        c.font = BAD_FONT
        c.fill = BAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    ws.cell(r, 1, "三、分数分布").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for label in ['强利多 (≥+7)', '中利多 (+4~+6)', '弱利多 (+1~+3)',
                  '弱利空 (-1~-3)', '中利空 (-4~-6)', '强利空 (≤-7)']:
        ws.cell(r, 2, label).font = H2_FONT
        cell = ws.cell(r, 3, score_dist.get(label, 0))
        cell.font = BOLD
        cell.alignment = Alignment(horizontal='center')
        r += 1

    r += 1
    ws.cell(r, 1, "四、短线交易参考").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for label, key, fill, font in [
        ('重点跟踪', 'focus_targets', GOOD_FILL, GOOD_FONT),
        ('风险回避', 'avoid_targets', BAD_FILL, BAD_FONT),
    ]:
        if short_term.get(key):
            ws.cell(r, 2, label).font = H2_FONT
            cell = ws.cell(r, 3, '; '.join(short_term[key]))
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.row_dimensions[r].height = 36
            r += 1
    # v9.30 (老板铁律): 删"资金偏好预判" 段, 改"短线交易提醒: 看公告到底多利好"
    ws.cell(r, 2, "短线交易提醒").font = H2_FONT
    c = ws.cell(r, 3, "仔细看公告到底多利好, 不要被板块/行业标签迷惑 — 同一行业可能利好利空参半, 关键是单个公告的具体内容和催化强度。利好需具体到金额/比例/对手方/期限, 利空需具体到减持比例/立案主体/减值规模。")
    c.font = NEUTRAL_FONT
    c.alignment = Alignment(wrap_text=True, vertical='center')
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 60
    r += 1

    _render_top5_sheet(wb, '每日全部利好', top_good, side='good')
    _render_top5_sheet(wb, '每日全部利空', top_bad, side='bad')
    # v9.34: 多空全表被"今日全部公告" 替代 (1391 条全铺开, watchlist 命中标黄)
    _render_excluded_sheet(wb, excluded, neutral_ann)
    # v9.30 (老板铁律): 删 _render_signals_sheet 调用 — "PDF 最后一部份的附录也可以删掉了" 同样适用 Excel
    # 空公告日也保留稳定的 5-sheet 契约；raw=None 仅表示调用方没有提供底稿。
    if raw is not None:
        _render_all_anns_sheet(wb, raw, data)

    wb.save(output_path)
    print(f"OK Excel: {output_path} ({len(wb.sheetnames)} sheets)")
    return output_path


def _render_top5_sheet(wb: Workbook, name: str, items: List[Dict], side: str) -> None:
    fill = GOOD_FILL if side == 'good' else BAD_FILL
    font = GOOD_FONT if side == 'good' else BAD_FONT
    link_font = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
    ws = wb.create_sheet(name)
    # 9.x: sheet tab 颜色跟数据方向一致 (A 股习惯: 利好红 / 利空绿), 一眼分辨
    tab_color = 'C00000' if side == 'good' else '2E7D32'
    ws.sheet_properties.tabColor = tab_color
    widths = [4, 8, 20, 12, 12, 30, 35, 22, 12, 50, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 1
    for i, h in enumerate(['#', '代码', '公司', '行业', '公告日', '事件', '一句话概括', '具体数字', '链接', '判断逻辑', '分数'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 11)
    r += 1
    for it in items:
        ws.cell(r, 1, it.get('rank', ''))
        ws.cell(r, 2, it.get('code', ''))
        ws.cell(r, 3, it.get('company', ''))
        ws.cell(r, 4, it.get('industry', ''))
        ws.cell(r, 5, it.get('best_date', ''))
        ws.cell(r, 6, it.get('event', ''))  # 事件(短, 不再放全标题以省列宽)
        # 一句话概括 (业务定性 + 关键数字)
        ws.cell(r, 7, it.get('best_summary', '') or '')
        # 关键数字: 换行 + wrap_text
        facts = it.get('best_facts', [])
        facts_str = '\n'.join(facts[:6]) if facts else '(公告未披露具体数字, 见下方公告链接)'
        ws.cell(r, 8, facts_str)
        # 链接(超链接形式)
        url = it.get('best_url', '')
        if url:
            link_cell = ws.cell(r, 9, '📄 打开公告 PDF')
            link_cell.hyperlink = url
            link_cell.font = link_font
        else:
            ws.cell(r, 9, '')
        ws.cell(r, 10, it.get('logic', ''))
        ws.cell(r, 11, it.get('best_score', ''))
        style_data(ws, r, 11, fill=fill, font=font)
        ws.row_dimensions[r].height = 80
        r += 1


def _render_companies_sheet(wb: Workbook, name: str, good: List[Dict], bad: List[Dict]) -> None:
    ws = wb.create_sheet(name)
    widths = [4, 8, 20, 12, 12, 50, 25, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 1
    ws.cell(r, 1, f"多空全表 (利好 {len(good)} | 利空 {len(bad)})").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2
    for header, items, fill, font in [
        ('【利好公司】', good, GOOD_FILL, GOOD_FONT),
        ('【利空公司】', bad, BAD_FILL, BAD_FONT),
    ]:
        ws.cell(r, 1, header).font = Font(name='Microsoft YaHei', size=12, bold=True, color='1F4E78')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for i, h in enumerate(['#', '代码', '公司', '行业', '公告日', '代表事件', '具体数字', '分数'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 8)
        r += 1
        for i, it in enumerate(items, 1):
            ws.cell(r, 1, i).alignment = Alignment(horizontal='center')
            ws.cell(r, 2, it.get('code', ''))
            ws.cell(r, 3, it.get('company', ''))
            ws.cell(r, 4, it.get('industry', ''))
            ws.cell(r, 5, it.get('best_date', ''))
            ws.cell(r, 6, it.get('best_title', ''))
            facts = it.get('best_facts', [])
            facts_str = '\n'.join(facts[:6]) if facts else ''
            ws.cell(r, 7, facts_str)
            ws.cell(r, 8, it.get('best_score', ''))
            style_data(ws, r, 8, fill=fill, font=font)
            ws.row_dimensions[r].height = 36
            r += 1
        r += 1


def _render_excluded_sheet(wb: Workbook, excluded: List, neutral_ann: List = None) -> None:
    """
    排除清单 sheet: 顶部显示中性原因汇总, 下方显示完整公告明细
    """
    ws = wb.create_sheet('每日全部中性')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50
    r = 1
    # === 顶部: 原因汇总 ===
    ws.cell(r, 1, "中性过滤原因汇总").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2
    for i, h in enumerate(['#', '中性原因'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 2)
    r += 1
    for i, reason in enumerate(excluded, 1):
        ws.cell(r, 1, i).alignment = Alignment(horizontal='center')
        ws.cell(r, 2, reason)
        style_data(ws, r, 2, fill=WARN_FILL, font=WARN_FONT)
        r += 1
    r += 2

    # === 下方: 中性公告明细 ===
    if neutral_ann:
        ws.cell(r, 1, f"中性公告明细 (共 {len(neutral_ann)} 条)").font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 2
        for i, h in enumerate(['#', '代码', '公司', '公告日', '公告标题', '中性原因', '链接'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 7)
        r += 1
        for i, n in enumerate(neutral_ann, 1):
            ws.cell(r, 1, i).alignment = Alignment(horizontal='center')
            ws.cell(r, 2, n.get('code', ''))
            ws.cell(r, 3, n.get('name', ''))
            ws.cell(r, 4, n.get('ann_date', ''))
            title_cell = ws.cell(r, 5, n.get('title', ''))
            title_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(r, 6, n.get('reason', ''))
            url = n.get('url', '')
            if url:
                link_cell = ws.cell(r, 7, f'=HYPERLINK("{url}","📄 原文")')
                link_cell.font = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
                link_cell.alignment = Alignment(vertical='top')
            # 数据行底色 (灰)
            for c in range(1, 8):
                ws.cell(r, c).fill = WARN_FILL
            ws.row_dimensions[r].height = 28
            r += 1





def _render_all_anns_sheet(wb: Workbook, raw_anns: List[Dict], processed: Dict) -> None:
    """v9.34 + 老板 6/18 二次拍: 今日全部公告 sheet 严格按 watchlist 筛选
    - raw_anns 入口就过滤 (v9.34.1 老板拍: outcome 必须严格按 watchlist 标的筛选, 不命中的不进 sheet)
    - 命中 YTD top 15 申万一级 watchlist 的标黄 (#FFF7CC)
    - 状态列: 利好 / 利空 / 中性(未入榜) / 不在watchlist
    - 行业列: watchlist 优先 (申万一级) + 兜底 infer_industry
    - 分数列: 命中信号的分, 未入榜为 0
    """
    import analyze
    # 加载 watchlist
    analyze._load_watchlist()
    wl_codes = analyze._WATCHLIST_CODES
    wl_industry = analyze._WATCHLIST_INDUSTRY

    # v9.34.1: 老板 6/18 拍 — outcome 严格按 watchlist 标的筛选, 不命中的不进 sheet
    _raw_total = len(raw_anns)
    raw_anns = [
        a for a in raw_anns
        if str(a.get('secCode', '')).strip() in wl_codes
        and not analyze.is_earnings_forecast_title(a.get('announcementTitle', '').strip())
    ]
    _filtered_out = _raw_total - len(raw_anns)

    # 按公告 ID 精确索引,避免同公司多公告时中性公告覆盖利好/利空 best item。
    # 例: 大族激光 2026-06-26 同日 5 条公告,其中扩产应为 +7,其它治理公告仍为中性。
    score_index = {}
    company_score_index = {}

    def _ann_id_from_url(url: str) -> str:
        m = re.search(r'/([0-9]+)\.PDF', url or '', re.I)
        return m.group(1) if m else ''

    for it in processed.get('top_good', []):
        # best_ann_id / best_code
        c = it.get('code', '')
        if c:
            row = ('利好', it.get('best_score', 0), it.get('industry', ''), it.get('best_url', ''), it.get('best_summary', ''))
            ann_id = _ann_id_from_url(it.get('best_url', ''))
            if ann_id:
                score_index[(c, ann_id)] = row
            company_score_index[(c, '')] = row
    for it in processed.get('all_bad_companies', []):
        c = it.get('code', '')
        if c:
            row = ('利空', it.get('best_score', 0), it.get('industry', ''), it.get('best_url', ''), it.get('best_summary', ''))
            ann_id = _ann_id_from_url(it.get('best_url', ''))
            if ann_id:
                score_index[(c, ann_id)] = row
            company_score_index[(c, '')] = row
    for it in processed.get('neutral_announcements', []):
        c = it.get('code', '')
        if c:
            ann_id = _ann_id_from_url(it.get('url', ''))
            if ann_id:
                score_index[(c, ann_id)] = ('中性', 0, '', it.get('url', ''), '')

    # 黄色 fill (watchlist 命中)
    YELLOW_FILL = PatternFill('solid', fgColor='FFF7CC')
    YELLOW_FONT = Font(name='Microsoft YaHei', size=9, color='1F4E78')
    NEU_GRAY_FONT = Font(name='Microsoft YaHei', size=9, color='9E9E9E')

    ws = wb.create_sheet('今日全部公告')
    ws.sheet_properties.tabColor = 'FFC000'  # 黄色 tab
    widths = [4, 8, 18, 12, 18, 50, 12, 8, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 1
    for i, h in enumerate(['#', '代码', '公司', '行业', '公告日', '事件', '状态', '分数', '链接'], 1):
        c = ws.cell(r, i, h)
        c.font = Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1A4D8F')
        c.alignment = Alignment(horizontal='center', vertical='center')
    r += 1

    in_wl = 0
    not_in_wl = 0
    link_font = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
    for i, a in enumerate(raw_anns, 1):
        code = str(a.get('secCode', '')).strip()
        name = (a.get('secName', '') or '').replace('*', '').replace(' ', '').strip()
        title = a.get('announcementTitle', '') or ''
        ann_id = str(a.get('announcementId', ''))
        url = f"http://static.cninfo.com.cn/{a.get('adjunctUrl', '')}"
        ann_time_ms = a.get('announcementTime', 0)
        ann_date = ''
        if ann_time_ms:
            try:
                from datetime import datetime
                ann_date = datetime.fromtimestamp(ann_time_ms / 1000).strftime('%Y-%m-%d %H:%M')
            except Exception:
                ann_date = ''

        # 行业: watchlist 优先
        if code in wl_codes:
            industry = wl_industry.get(code, '')
        else:
            industry = analyze.infer_industry(name, code) or ''

        # 状态 / 分数: 优先公告 ID 精确匹配,再回退公司级 best item。
        status, score, _, _, summary = score_index.get(
            (code, ann_id),
            company_score_index.get((code, ''), ('', 0, '', '', '')),
        )
        if not status:
            status = '中性 (未入榜)'

        # v9.34.1: raw_anns 已按 watchlist 过滤 (函数开头), 这里 is_in_watchlist 永远 True
        is_in_watchlist = True
        in_wl += 1

        ws.cell(r, 1, i)
        ws.cell(r, 2, code)
        ws.cell(r, 3, name)
        ws.cell(r, 4, industry)
        ws.cell(r, 5, ann_date)
        ws.cell(r, 6, title)
        ws.cell(r, 7, status)
        ws.cell(r, 8, score)
        if url:
            link_cell = ws.cell(r, 9, '📄 打开公告 PDF')
            link_cell.hyperlink = url
            link_cell.font = link_font
        # 格式化 + 标黄 (watchlist 命中)
        for col in range(1, 10):
            cell = ws.cell(r, col)
            cell.border = THIN
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if is_in_watchlist:
                cell.fill = YELLOW_FILL
                if not cell.font or not cell.font.color:
                    cell.font = YELLOW_FONT
            else:
                if not cell.font or not cell.font.color:
                    cell.font = NEU_GRAY_FONT
        # 状态颜色 (跟 利好红 / 利空绿 / 中性灰 一致)
        st_cell = ws.cell(r, 7)
        if status == '利好':
            st_cell.font = Font(name='Microsoft YaHei', size=9, color='C00000', bold=True)
        elif status == '利空':
            st_cell.font = Font(name='Microsoft YaHei', size=9, color='2E7D32', bold=True)
        else:
            st_cell.font = Font(name='Microsoft YaHei', size=9, color='9E9E9E')
        ws.row_dimensions[r].height = 22
        r += 1

    # v9.34.1: 老板 6/18 拍 — outcome 严格按 watchlist 筛选
    # _filtered_out 已经在函数开头过滤掉了, _render_all_anns_sheet 输出 100% 都是 watchlist 命中
    print(f'   今日全部公告 sheet: 总 {len(raw_anns)} 条 (raw {_raw_total} -> watchlist 命中 {len(raw_anns)}, 过滤掉 {_filtered_out} 条非命中)')


def _render_recap_sheet(wb: Workbook, recap: Dict) -> None:
    """今日复盘 sheet: 1.1 强度分档+综合命中率 / 1.2 利好全量 / 1.3 利空全量 / 1.4 复盘反思
    (T 日 16:01 收盘后跑, 复盘 T 日报告在 T 日的市场反应, 涨跌幅 = T 日前一交易日 close -> T 日 close)
    """
    """今日复盘 sheet: T 日 16:01 收盘后跑, 复盘 T 日报告在 T 日的市场反应"""
    if not recap:
        return
    prev_date = recap.get("prev_date", "?")
    source_entry_date = recap.get("source_entry_date") or prev_date
    target_trade_date = recap.get("target_trade_date") or prev_date
    kline_avail = recap.get("kline_avail", False)
    good_items = recap.get("good_items", []) or []
    bad_items = recap.get("bad_items", []) or []
    good_stats = recap.get("good_stats", {}).get("buckets", [])
    bad_stats = recap.get("bad_stats", {}).get("buckets", [])
    market_ctx = recap.get("market_ctx") or {}  # v9.31-C: 噪音日反向档位 FAIL 角标

    ws = wb.create_sheet('今日复盘')
    ws.column_dimensions['A'].width = 4
    for i, w in enumerate([6, 10, 20, 16, 14, 10, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    r = 1
    kline_tag = "已收盘" if kline_avail else "数据缺失 (T 日 K 线未生成)"
    ws.cell(r, 1, f"今日复盘 (入榜日={source_entry_date} -> 验证日={target_trade_date})").font = Font(
        name='Microsoft YaHei', size=14, bold=True, color='1F4E78')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(r, 1, f"复盘对象: {source_entry_date} 入榜标的  ·  涨跌幅口径: {target_trade_date} 前收 -> 收盘; 利好收红算命中, 利空收绿算命中  ·  K 线状态: {kline_tag}").font = Font(
        name='Microsoft YaHei', size=10, italic=True, color='808080')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    # v9.31: 复盘 sheet 头部加市场背景 (从 K 线 cache 反推 T-1 -> T 全市场均值)
    market_ctx = recap.get("market_ctx") or {}
    if market_ctx and market_ctx.get("avg_pct") is not None:
        avg = market_ctx["avg_pct"]
        n = market_ctx.get("sample_n", 0)
        label = market_ctx.get("label", "")
        note = market_ctx.get("note", "")
        t_minus_1 = market_ctx.get("t_minus_1", "?")
        t_day = market_ctx.get("prev_date", "?")
        ctx_font = PCT_UP_FONT if avg > 0.5 else (PCT_DOWN_FONT if avg < -0.5 else NEUTRAL_FONT)
        ws.cell(r, 1, f"入榜样本背景: {t_minus_1} → {t_day} 样本均值 {avg:+.2f}% (n={n})  ·  {label}  ·  {note}").font = ctx_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    r += 1

    # === 5.1 强度分档命中统计 ===
    ws.cell(r, 1, "1.1 强度分档命中统计").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for i, h in enumerate(['强度档', '数量', '命中数', '命中率', '平均涨跌幅'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 5)
    r += 1
    for label, stats, fill, font, side in [
        ('【利好命中】', good_stats, GOOD_FILL, GOOD_FONT, "good"),
        ('【利空命中】', bad_stats, BAD_FILL, BAD_FONT, "bad"),
    ]:
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 1).fill = fill
        r += 1
        for b in stats:
            strength_cell = b["strength"]
            # v9.31-C: 噪音日反向档位 FAIL 加角标 [普涨压制] / [普跌压制]
            mlabel = market_ctx.get("label", "") if market_ctx else ""
            if b.get("count", 0) > 0 and b.get("hit_rate", 0) < 50:
                if "普涨" in mlabel and side == "bad":
                    strength_cell = f"{b['strength']} [普涨压制]"
                elif "普跌" in mlabel and side == "good":
                    strength_cell = f"{b['strength']} [普跌压制]"
            ws.cell(r, 1, strength_cell).font = font
            ws.cell(r, 1).fill = fill
            ws.cell(r, 2, b["count"])
            ws.cell(r, 3, b["hit"])
            rate = f"{b['hit_rate']:.1f}%" if b.get("count") else "-"
            avg = f"{b['avg_pct']:+.2f}%" if b.get("count") else "-"
            ws.cell(r, 4, rate)
            ws.cell(r, 5, avg)
            style_data(ws, r, 5, fill=fill, font=font)
            r += 1
        r += 1

    def _render_items(title, items, fill, font):
        nonlocal r
        ws.cell(r, 1, f"{title} ({len(items)} 只)").font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for i, h in enumerate(['#', '代码', '公司', '行业', '强度', '分数', '涨跌幅', '状态'], 1):
            ws.cell(r, i, h)
        style_header(ws, r, 8)
        r += 1
        for i, it in enumerate(items, 1):
            pct = it.get("pct")
            pct_str = f"{pct:+.2f}%" if pct is not None else "-"
            tag = it.get("tag", "")
            tag_pstyle = WARN_FILL if tag in ("涨停", "跌停") else None
            ws.cell(r, 1, i).alignment = Alignment(horizontal='center')
            ws.cell(r, 2, it.get("code", ""))
            ws.cell(r, 3, it.get("company", ""))
            ws.cell(r, 4, it.get("industry", ""))
            ws.cell(r, 5, it.get("strength", ""))
            ws.cell(r, 6, f"{it.get('score', 0):+d}").alignment = Alignment(horizontal='center')
            pct_cell = ws.cell(r, 7, pct_str)
            pct_cell.alignment = Alignment(horizontal='center')
            if pct is not None:
                # v9.2: 涨跌颜色按方向 (A 股习惯: 涨红跌绿), 不用 side 颜色
                if pct > 0:
                    pct_cell.font = PCT_UP_FONT
                    pct_cell.fill = PCT_UP_FILL
                elif pct < 0:
                    pct_cell.font = PCT_DOWN_FONT
                    pct_cell.fill = PCT_DOWN_FILL
                else:
                    pct_cell.font = font
                    pct_cell.fill = fill
            ws.cell(r, 8, tag).alignment = Alignment(horizontal='center')
            style_data(ws, r, 8, fill=fill, font=font, skip_cols=[7])
            r += 1
        r += 1

    _render_items("1.2 利好全量明细", good_items, GOOD_FILL, GOOD_FONT)
    _render_items("1.3 利空全量明细", bad_items, BAD_FILL, BAD_FONT)

    # === 1.1 综合命中率 (合并到 1.1 段尾) ===
    ws.cell(r, 1, "1.1 综合命中率 (按方向)").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    for side, stats in [("利好", good_stats), ("利空", bad_stats)]:
        total = sum(b.get("count", 0) for b in stats)
        hit = sum(b.get("hit", 0) for b in stats)
        if total == 0:
            rate_str = "数据不足 (K 线缺失, 待 7am 跑批回填)"
        else:
            rate = round(hit / total * 100, 1)
            rate_str = f"{hit}/{total} = {rate}%"
        ws.cell(r, 1, f"{side}综合命中率").font = BOLD
        ws.cell(r, 2, rate_str)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        r += 1
    r += 1

    ws.cell(r, 1, "1.4 复盘反思和改进方向").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    if not good_stats and not bad_stats:
        ws.cell(r, 1, "数据不足, 暂不生成反思 (K 线缺失或 T 日前一交易日报告无入榜标的)")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    else:
        from collections import Counter
        # ===== A. 强误判样本清单 (利好反跌, 表格) =====
        misjudged = [it for it in (good_items or []) if it.get("score", 0) >= 4 and (it.get("pct") or 0) <= -2.0]
        misjudged.sort(key=lambda x: x.get("pct") or 0)
        ws.cell(r, 1, f"A. 强误判样本清单 (score≥4 且涨跌幅≤-2%, {len(misjudged)} 只)").font = H2_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        if not misjudged:
            ws.cell(r, 1, "本交易日无强误判样本, 强利多方向判断准确")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        else:
            for i, h in enumerate(["#", "代码", "公司", "行业", "主信号", "分数", "涨跌幅", "事件"], 1):
                c = ws.cell(r, i, h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
            style_header(ws, r, 8)
            r += 1
            for i, it in enumerate(misjudged[:10], 1):
                pct = it.get("pct")
                pct_str = f"{pct:+.2f}%" if pct is not None else "-"
                row_vals = [
                    i,
                    it.get("code", ""),
                    it.get("company", ""),
                    it.get("industry", "") or "-",
                    it.get("primary_signal") or "-",
                    f"{it.get('score', 0):+d}",
                    pct_str,
                    (it.get("event") or "")[:40],
                ]
                for j, v in enumerate(row_vals, 1):
                    cell = ws.cell(r, j, v)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if j in (5, 6):
                        cell.font = GOOD_FONT  # 主信号 / 分数 (Excel 利好配色)
                    elif j == 7:
                        pct_val = it.get("pct")
                        if pct_val is not None:
                            if pct_val > 0:
                                cell.font = PCT_UP_FONT
                                cell.fill = PCT_UP_FILL
                            elif pct_val < 0:
                                cell.font = PCT_DOWN_FONT
                                cell.fill = PCT_DOWN_FILL
                            else:
                                cell.font = NEUTRAL_FONT
                        else:
                            cell.font = NEUTRAL_FONT
                    else:
                        cell.font = NEUTRAL_FONT
                r += 1
        r += 1

        # ===== B. 强反向样本清单 (利空反涨, 表格, v9.5 新增) =====
        reverse_strong = [it for it in (bad_items or []) if it.get("score", 0) <= -4 and (it.get("pct") or 0) >= 2.0]
        reverse_strong.sort(key=lambda x: -(x.get("pct") or 0))
        ws.cell(r, 1, f"B. 强反向样本清单 (score≤-4 且涨跌幅≥+2%, {len(reverse_strong)} 只)").font = H2_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        if not reverse_strong:
            ws.cell(r, 1, "本交易日无强反向样本, 强利空方向判断准确")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        else:
            for i, h in enumerate(["#", "代码", "公司", "行业", "主信号", "分数", "涨跌幅", "事件"], 1):
                c = ws.cell(r, i, h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
            style_header(ws, r, 8)
            r += 1
            for i, it in enumerate(reverse_strong[:10], 1):
                pct = it.get("pct")
                pct_str = f"{pct:+.2f}%" if pct is not None else "-"
                row_vals = [
                    i,
                    it.get("code", ""),
                    it.get("company", ""),
                    it.get("industry", "") or "-",
                    it.get("primary_signal") or "-",
                    f"{it.get('score', 0):+d}",
                    pct_str,
                    (it.get("event") or "")[:40],
                ]
                for j, v in enumerate(row_vals, 1):
                    cell = ws.cell(r, j, v)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if j in (5, 6):
                        cell.font = BAD_FONT  # 利空方向: 主信号/分数红字 (与利好反向)
                    elif j == 7:
                        pct_val = it.get("pct")
                        if pct_val is not None:
                            if pct_val > 0:
                                cell.font = PCT_UP_FONT
                                cell.fill = PCT_UP_FILL
                            elif pct_val < 0:
                                cell.font = PCT_DOWN_FONT
                                cell.fill = PCT_DOWN_FILL
                            else:
                                cell.font = NEUTRAL_FONT
                        else:
                            cell.font = NEUTRAL_FONT
                    else:
                        cell.font = NEUTRAL_FONT
                r += 1
        r += 1

        # ===== 1.4 C. 全部利好误判 (NEW 全量表) =====
        all_good_miss = [it for it in (good_items or []) if (it.get("pct") or 0) <= 0]
        all_good_miss.sort(key=lambda x: (-x.get("score", 0), x.get("pct") or 0))
        sg = sum(1 for it in all_good_miss if it.get("strength") == "强利多")
        mg = sum(1 for it in all_good_miss if it.get("strength") == "中利多")
        wg = sum(1 for it in all_good_miss if it.get("strength") == "弱利多")
        ws.cell(r, 1, f"C. 全部利好误判 (推荐利好但实际跌/平, 共 {len(all_good_miss)} 只, 强利多 {sg} / 中利多 {mg} / 弱利多 {wg})").font = H2_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        if not all_good_miss:
            ws.cell(r, 1, "本交易日无利好误判, 利好方向判断准确")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        else:
            for i, h in enumerate(["#", "代码", "公司", "行业", "主信号", "分数", "涨跌幅", "事件"], 1):
                c = ws.cell(r, i, h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
            style_header(ws, r, 8)
            r += 1
            for i, it in enumerate(all_good_miss, 1):  # 全量, 不限 10
                pct = it.get("pct")
                pct_str = f"{pct:+.2f}%" if pct is not None else "-"
                row_vals = [
                    i, it.get("code", ""), it.get("company", ""),
                    it.get("industry", "") or "-",
                    it.get("primary_signal") or "-",
                    f"{it.get('score', 0):+d}",
                    pct_str,
                    (it.get("event") or "")[:40],
                ]
                for j, v in enumerate(row_vals, 1):
                    cell = ws.cell(r, j, v)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if j in (5, 6):
                        cell.font = GOOD_FONT
                    elif j == 7:
                        pct_val = it.get("pct")
                        if pct_val is not None:
                            if pct_val > 0:
                                cell.font = PCT_UP_FONT
                                cell.fill = PCT_UP_FILL
                            elif pct_val < 0:
                                cell.font = PCT_DOWN_FONT
                                cell.fill = PCT_DOWN_FILL
                            else:
                                cell.font = NEUTRAL_FONT
                        else:
                            cell.font = NEUTRAL_FONT
                    else:
                        cell.font = NEUTRAL_FONT
                r += 1
        r += 1

        # ===== 1.4 D. 全部利空误判 (NEW 全量表) =====
        all_bad_miss = [it for it in (bad_items or []) if (it.get("pct") or 0) >= 0]
        all_bad_miss.sort(key=lambda x: (x.get("score", 0), -(x.get("pct") or 0)))
        sb = sum(1 for it in all_bad_miss if it.get("strength") == "强利空")
        wb_n = sum(1 for it in all_bad_miss if it.get("strength") == "弱利空")
        ws.cell(r, 1, f"D. 全部利空误判 (推荐利空但实际涨/平, 共 {len(all_bad_miss)} 只, 强利空 {sb} / 弱利空 {wb_n})").font = H2_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        if not all_bad_miss:
            ws.cell(r, 1, "本交易日无利空反向, 利空方向判断准确")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        else:
            for i, h in enumerate(["#", "代码", "公司", "行业", "主信号", "分数", "涨跌幅", "事件"], 1):
                c = ws.cell(r, i, h)
                c.font = HEADER_FONT
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal='center', vertical='center')
            style_header(ws, r, 8)
            r += 1
            for i, it in enumerate(all_bad_miss, 1):  # 全量
                pct = it.get("pct")
                pct_str = f"{pct:+.2f}%" if pct is not None else "-"
                row_vals = [
                    i, it.get("code", ""), it.get("company", ""),
                    it.get("industry", "") or "-",
                    it.get("primary_signal") or "-",
                    f"{it.get('score', 0):+d}",
                    pct_str,
                    (it.get("event") or "")[:40],
                ]
                for j, v in enumerate(row_vals, 1):
                    cell = ws.cell(r, j, v)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if j in (5, 6):
                        cell.font = BAD_FONT  # 利空配色
                    elif j == 7:
                        pct_val = it.get("pct")
                        if pct_val is not None:
                            if pct_val > 0:
                                cell.font = PCT_UP_FONT
                                cell.fill = PCT_UP_FILL
                            elif pct_val < 0:
                                cell.font = PCT_DOWN_FONT
                                cell.fill = PCT_DOWN_FILL
                            else:
                                cell.font = NEUTRAL_FONT
                        else:
                            cell.font = NEUTRAL_FONT
                    else:
                        cell.font = NEUTRAL_FONT
                r += 1
        r += 1

        # ===== 1.5 误判原因分析 (根因 4 类) =====
        ws.cell(r, 1, "1.5 误判原因分析 (根因)").font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        # 复用 build_pdf 的根因分析函数
        import sys as _sys
        if "build_pdf" not in _sys.modules:
            import build_pdf as _bp
        else:
            _bp = _sys.modules["build_pdf"]
        for ln in _bp._build_root_causes_pdf(good_items, bad_items, good_stats, bad_stats, all_good_miss, all_bad_miss):
            ws.cell(r, 1, ln)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
        r += 1

        # ===== 1.6 改进方向 (actionable 5 条) =====
        ws.cell(r, 1, "1.6 改进方向 (actionable)").font = TITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for ln in _bp._build_action_items_pdf(good_items, bad_items, good_stats, bad_stats, all_good_miss, all_bad_miss):
            ws.cell(r, 1, ln)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1

def render_recap_workbook(processed_json_path: Path, output_xlsx_path: Path) -> Path:
    """v9.x 复盘独立 Excel: 只渲染 data['recap'] 段
    数据契约: processed.json 含 'recap' 字段 (run.py Step 2.5 写入)
    """
    with open(processed_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    recap = data.get("recap") or {}
    if not recap:
        print(f"[WARN] {processed_json_path} 不含 recap 段, 复盘 Excel 跳过")
        return output_xlsx_path
    wb = Workbook()
    wb.remove(wb.active)
    _render_recap_sheet(wb, recap)
    wb.save(output_xlsx_path)
    print(f"OK 复盘 Excel: {output_xlsx_path} (1 sheet)")
    return output_xlsx_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 build_excel.py <processed.json> <output.xlsx> [--recap]")
        sys.exit(1)
    if "--recap" in sys.argv:
        sys.argv.remove("--recap")
        render_recap_workbook(Path(sys.argv[1]), Path(sys.argv[2]))
    else:
        render(Path(sys.argv[1]), Path(sys.argv[2]))


def render_weekly_workbook(weekly_data: Dict, output_xlsx_path: Path) -> Path:
    """周报 Excel: 多日聚合 6 档命中率 + 趋势 + 铁律提示 (Sunday 20:00 自动跑)

    数据契约: weekly_data = {
      'days': [...],
      'agg': {'good': [...], 'bad': [...]},
      'summary': {good_hit, good_total, good_rate, bad_hit, bad_total, bad_rate},
      'trend': [str, ...],
      'abnormal_locked': str|None
    }
    """
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('周报')
    ws.column_dimensions['A'].width = 4
    for i, w in enumerate([18, 10, 10, 12, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    days = weekly_data["days"]
    r = 1
    ws.cell(r, 1, f"巨潮资讯 周报 ({days[0]} ~ {days[-1]})").font = Font(
        name='Microsoft YaHei', size=14, bold=True, color='1F4E78')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, f"覆盖区间: {days[0]} ~ {days[-1]} ({len(days)} 个交易日) · 口径: 多日 6 档滚动命中率 · 生成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(
        name='Microsoft YaHei', size=9, italic=True, color='808080')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    # === 一、综合命中率 ===
    s = weekly_data["summary"]
    ws.cell(r, 1, "一、综合命中率 (跨 3 档)").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for i, h in enumerate(['方向', '命中', '样本', '命中率'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 4)
    r += 1
    rows_summary = [
        ('利好 (跨 3 档)', s['good_hit'], s['good_total'], s['good_rate'], GOOD_FILL, GOOD_FONT),
        ('利空 (跨 3 档)', s['bad_hit'], s['bad_total'], s['bad_rate'], BAD_FILL, BAD_FONT),
    ]
    for label, hit, total, rate, fill, font in rows_summary:
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 1).fill = fill
        ws.cell(r, 2, hit)
        ws.cell(r, 3, total)
        rate_cell = ws.cell(r, 4, f"{rate:.1f}%")
        rate_cell.font = font
        rate_cell.fill = fill
        rate_cell.alignment = Alignment(horizontal='center')
        for c in range(2, 5):
            ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
        r += 1
    r += 1

    # === 二、6 档分档 ===
    ws.cell(r, 1, "二、6 档分档滚动命中率").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for i, h in enumerate(['档位', '命中', '样本', '命中率', '平均涨跌幅', '判定'], 1):
        ws.cell(r, i, h)
    style_header(ws, r, 6)
    r += 1
    for side, label in [('good', '利好'), ('bad', '利空')]:
        for b in weekly_data['agg'][side]:
            fill = GOOD_FILL if side == 'good' else BAD_FILL
            font = GOOD_FONT if side == 'good' else BAD_FONT
            ws.cell(r, 1, f"{b['strength']}({label})").font = font
            ws.cell(r, 1).fill = fill
            ws.cell(r, 2, b['hit'])
            ws.cell(r, 3, b['count'])
            rate_str = f"{b['hit_rate']:.1f}%" if b.get('count') else '-'
            avg_str = f"{b['avg_pct']:+.2f}%" if b.get('count') else '-'
            ws.cell(r, 4, rate_str)
            ws.cell(r, 5, avg_str)
            if b['count'] == 0:
                judge = "N/A (0 样本)"
            elif b['hit_rate'] >= 50:
                judge = "达标"
            else:
                judge = f"miss ({b['hit_rate']:.1f}%)"
            judge_cell = ws.cell(r, 6, judge)
            judge_cell.alignment = Alignment(horizontal='center')
            for c in range(2, 7):
                ws.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
            style_data(ws, r, 6, fill=fill, font=font, align='center')
            r += 1
    r += 1

    # === 三、趋势分析 ===
    ws.cell(r, 1, "三、趋势分析 (7 日滚动)").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    trend = weekly_data.get('trend', [])
    if not trend:
        ws.cell(r, 1, "样本不足, 暂不生成")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    else:
        for line in trend:
            ws.cell(r, 1, line)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1
    r += 1

    # === 四、铁律提示 ===
    ws.cell(r, 1, "四、铁律提示").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    iron_rules = [
        "铁律 #1: 异动类拆档: 普通异常波动 -2 / 严重异常波动 -4 / 严重异常波动+风险提示 -5",
        "铁律 #2: 单日噪音不改分, 复盘报告加说明",
        "铁律 #3: 0 样本 N/A, 不要求 50%",
        "铁律 #4: 档位 1-3 弱 / 4-6 中 / 7-10 强",
        "铁律 #5: 多公告全显示, 不取舍不遗漏",
        "铁律 #6: 用户说 '请停止' 立即停",
    ]
    for ir in iron_rules:
        ws.cell(r, 1, f"● {ir}")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    if weekly_data.get('abnormal_locked'):
        ws.cell(r, 1, f"! {weekly_data['abnormal_locked']}").font = BAD_FONT
        ws.cell(r, 1).fill = BAD_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1
    r += 1

    # === 五、覆盖日期 ===
    ws.cell(r, 1, "五、覆盖日期").font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, "、".join(days))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    wb.save(output_xlsx_path)
    print(f"OK weekly Excel: {output_xlsx_path} (1 sheet)")
    return output_xlsx_path
