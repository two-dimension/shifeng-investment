import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import analyze
import build_excel
import recap


class _FakeResponse:
    def __init__(self, payload):
        self._payload = payload

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        return False

    def read(self):
        return self._payload


class WeekAuditRegressionTests(unittest.TestCase):
    def test_empty_raw_still_creates_all_announcements_sheet(self):
        data = {
            "date": "2026-08-09",
            "coverage": {
                "start_date": "2026-08-09",
                "end_date": "2026-08-09",
                "range_label": "2026-08-09",
                "day_count": 1,
                "actual_dates": [],
            },
            "per_day": {"2026-08-09": 0},
            "fetch_meta": {"total": 0},
            "sentiment": {"good_count": 0, "bad_count": 0, "neutral_count": 0},
            "top_good": [],
            "top_bad": [],
            "all_good_companies": [],
            "all_bad_companies": [],
            "neutral_announcements": [],
            "excluded_summary": [],
            "score_distribution": {},
            "short_term": {},
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            build_excel.render(data, path, raw=[])
            wb = load_workbook(path, read_only=True)
            self.assertEqual(
                wb.sheetnames,
                ["Brief", "每日全部利好", "每日全部利空", "每日全部中性", "今日全部公告"],
            )
            self.assertEqual(wb["今日全部公告"].max_row, 1)
            wb.close()

    def test_market_context_is_deduplicated_and_has_prior_trade_date(self):
        items = [
            {"code": "000001", "pct": 2.0},
            {"code": "000001", "pct": 9.0},
            {"code": "600000", "pct": -1.0},
            {"code": "300001", "pct": None},
        ]
        result = recap._compute_market_ctx("2026-08-10", items)
        self.assertEqual(result["t_minus_1"], "2026-08-07")
        self.assertEqual(result["sample_n"], 2)
        self.assertEqual(result["avg_pct"], 0.5)

    def test_sina_fallback_respects_historical_as_of_date(self):
        rows = [
            {"day": "2026-07-30", "close": "10"},
            {"day": "2026-07-31", "close": "11"},
            {"day": "2026-08-03", "close": "12"},
            {"day": "2026-08-04", "close": "13"},
            {"day": "2026-08-05", "close": "14"},
            {"day": "2026-08-06", "close": "99"},
        ]
        payload = ("var=(" + json.dumps(rows) + ")").encode("gbk")
        with tempfile.TemporaryDirectory() as tmpdir, patch.object(
            analyze, "PRICE_CACHE", Path(tmpdir)
        ), patch.object(
            analyze, "_HAVE_AKSHARE_PRICE", False
        ), patch.object(
            analyze.urllib.request, "urlopen", return_value=_FakeResponse(payload)
        ):
            result = analyze._get_price_position("600000", as_of_date="2026-08-05")
        self.assertEqual(result["as_of_date"], "2026-08-05")
        self.assertEqual(result["current"], 14.0)
        self.assertEqual(result["high_60d"], 14.0)
        self.assertNotEqual(result["current"], 99.0)


if __name__ == "__main__":
    unittest.main()
