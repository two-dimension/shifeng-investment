#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""风险提示 Excel 输出。"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_FONT = Font(name="Microsoft YaHei", size=15, bold=True, color="FFFFFF")
H2_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="7A1E1E")
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
RED_FILL = PatternFill("solid", fgColor="C00000")
LIGHT_RED_FILL = PatternFill("solid", fgColor="FCE4D6")
ORANGE_FILL = PatternFill("solid", fgColor="F4B183")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")
DARK_FILL = PatternFill("solid", fgColor="7A1E1E")


def _load(data_or_path):
    if isinstance(data_or_path, (str, Path)):
        return json.loads(Path(data_or_path).read_text(encoding="utf-8"))
    return data_or_path


def _style_header(ws, row: int, fill=DARK_FILL):
    for cell in ws[row]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws):
    for col in ws.columns:
        max_len = 8
        letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(val), 60))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row > 1:
                cell.font = BODY_FONT
        ws.column_dimensions[letter].width = max_len + 2


def _risk_rows(items: List[dict]) -> List[List[object]]:
    rows = []
    for it in items:
        rows.append([
            it.get("rank", ""),
            it.get("code", ""),
            it.get("company", ""),
            it.get("risk_level", ""),
            it.get("score", ""),
            it.get("primary_signal", ""),
            it.get("concept", ""),
            it.get("summary", ""),
            it.get("net_profit_text", ""),
            it.get("profit_extract_status", ""),
            it.get("title", ""),
            it.get("url", ""),
        ])
    return rows


def _add_risk_sheet(wb: Workbook, name: str, items: List[dict]):
    ws = wb.create_sheet(name)
    headers = ["#", "代码", "公司", "风险等级", "分数", "风险类型", "概念/指数", "一句话", "归母净利润(万元)", "取数状态", "公告标题", "链接"]
    ws.append(headers)
    _style_header(ws, 1)
    for row in _risk_rows(items):
        ws.append(row)
    for r in range(2, ws.max_row + 1):
        score = ws.cell(r, 5).value or 0
        fill = LIGHT_RED_FILL if score <= -7 else ORANGE_FILL if score <= -5 else GRAY_FILL
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill
    ws.freeze_panes = "A2"
    _autosize(ws)


def render(data_or_path, out_path: Path) -> Path:
    data = _load(data_or_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "风险概览"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"风险提示-公告扫描-{data.get('date', '')}"
    ws["A1"].fill = RED_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    cov = data.get("coverage", {})
    sent = data.get("sentiment", {})
    rows = [
        ("扫描区间", cov.get("range_label", "")),
        ("原始公告数", cov.get("raw_total", 0)),
        ("watchlist公告数", cov.get("watchlist_ann_count", 0)),
        ("风险条数", sent.get("risk_count", 0)),
        ("重大/高风险公司数", sent.get("major_risk_company_count", 0)),
        ("观察项条数", sent.get("observation_count", 0)),
        ("生成时间", data.get("generated_at", "")),
    ]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = H2_FONT
        ws.cell(r, 2, v)
        r += 1

    r += 1
    ws.cell(r, 1, "风险等级分布").font = H2_FONT
    r += 1
    ws.append(["风险等级", "数量"])
    _style_header(ws, r)
    for k, v in sent.get("by_level", {}).items():
        ws.append([k, v])

    r = ws.max_row + 2
    ws.cell(r, 1, "TOP风险类型").font = H2_FONT
    r += 1
    ws.append(["风险类型", "数量"])
    _style_header(ws, r)
    for k, v in sent.get("by_signal", {}).items():
        ws.append([k, v])
    _autosize(ws)

    _add_risk_sheet(wb, "重大风险", data.get("major_risks", []))
    _add_risk_sheet(wb, "全部风险明细", data.get("risks", []))
    _add_risk_sheet(wb, "观察项", data.get("observations", []))

    ws_ann = wb.create_sheet("今日全部公告_watchlist")
    ws_ann.append(["代码", "公司", "概念/指数", "归母净利润(万元)", "取数状态", "公告标题", "链接"])
    _style_header(ws_ann, 1)
    for it in data.get("watchlist_announcements", []):
        ws_ann.append([
            it.get("code"),
            it.get("company"),
            it.get("concept"),
            it.get("net_profit_text", ""),
            it.get("profit_extract_status", ""),
            it.get("title"),
            it.get("url"),
        ])
    ws_ann.freeze_panes = "A2"
    _autosize(ws_ann)

    ws_perf = wb.create_sheet("业绩预告_全部")
    ws_perf.append(["代码", "公司", "归母净利润(万元)", "取数来源", "取数状态", "公告标题", "链接"])
    _style_header(ws_perf, 1)
    for it in data.get("performance_forecasts", []):
        ws_perf.append([
            it.get("code"),
            it.get("company"),
            it.get("net_profit_text", ""),
            it.get("net_profit_source", ""),
            it.get("profit_extract_status", ""),
            it.get("title"),
            it.get("url"),
        ])
    ws_perf.freeze_panes = "A2"
    _autosize(ws_perf)

    ws_rules = wb.create_sheet("规则说明")
    ws_rules.append(["项目", "说明"])
    _style_header(ws_rules, 1)
    ws_rules.append(["PDF阈值", "仅展示分数 <= -7 的重大/高风险"])
    ws_rules.append(["Excel口径", "保留 -2 至 -10 全部风险项"])
    ws_rules.append(["watchlist", data.get("rules", {}).get("watchlist_path", "")])
    ws_rules.append(["免责声明", "本报告基于公开公告自动生成, 仅供内部研究使用, 不代表投资建议。"])
    _autosize(ws_rules)

    wb.save(out_path)
    return out_path


def render_recap_workbook(data_or_path, out_path: Path) -> Path:
    data = _load(data_or_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "风险复盘"
    recap = data.get("recap", {})
    ws.append([f"风险提示-今日复盘-{recap.get('date', data.get('date', ''))}"])
    ws["A1"].fill = RED_FILL
    ws["A1"].font = TITLE_FONT
    ws.append(["风险报告日", data.get("date", "")])
    ws.append(["复盘交易日", recap.get("date", "")])
    ws.append(["样本数", recap.get("count", 0)])
    ws.append(["平均涨跌幅", recap.get("avg_pct", "")])
    ws.append([])
    ws.append(["代码", "公司", "风险等级", "分数", "风险类型", "开盘", "收盘", "涨跌幅%", "公告标题"])
    _style_header(ws, ws.max_row)
    for it in recap.get("items", []):
        ws.append([
            it.get("code"), it.get("company"), it.get("risk_level"), it.get("score"),
            it.get("primary_signal"), it.get("open"), it.get("close"), it.get("pct"),
            it.get("title"),
        ])
    _autosize(ws)
    wb.save(out_path)
    return out_path
