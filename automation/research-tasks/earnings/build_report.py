#!/usr/bin/env python3
"""
A 股业绩预告日报生成脚本
入参：<input.json> <output_dir>
入参 JSON 结构：
{
  "date": "YYYY-MM-DD",
  "weekday": "周一",
  "items": [
    {
      "公告日期": "...", "证券代码": "...", "证券简称": "...",
      "预告类型": "预增|续盈|扭亏|预减|首亏|续亏|不确定",
      "预告期间": "...", "口径": "归母净利润|扣非净利润|...",
      "下限万元": 0.0, "上限万元": 0.0, "上年同期万元": 0.0,
      "同比下限%": 0.0, "同比上限%": 0.0,
      "原因摘要": "...",
      "AI主题相关": 0|1, "AI说明": "..."  (可选)
    }
  ]
}
输出：<output_dir>/A股业绩预告-<date>.xlsx 和 .pdf
"""
import sys
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.units import cm

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from industry_map import enrich_item_subset
from validate_input import check_item


# 字体: macOS 自带 STHeiti Medium.ttc,覆盖 Helvetica 的 CJK 缺失
def _register_fonts() -> tuple:
    candidates = [
        os.environ.get("RESEARCH_CJK_FONT"),
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/System/Library/Fonts/PingFang.ttc",
    ]
    font_path = next((Path(value) for value in candidates if value and Path(value).is_file()), None)
    if font_path is None:
        raise FileNotFoundError(
            "未找到中文 PDF 字体；请安装 fonts-noto-cjk 或设置 RESEARCH_CJK_FONT"
        )
    pdfmetrics.registerFont(TTFont("ResearchCJK", str(font_path)))
    return "ResearchCJK", "ResearchCJK"


FONT_REG, FONT_BOLD = _register_fonts()

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="305496")
AI_FILL = PatternFill("solid", fgColor="FFF2CC")

POSITIVE_TYPES = {"预增", "续盈", "扭亏"}
NEGATIVE_TYPES = {"预减", "首亏", "续亏"}
PDF_POSITIVE_COLOR = "#C00000"
PDF_NEGATIVE_COLOR = "#00A651"


def _pdf_name_color(forecast_type):
    if forecast_type in POSITIVE_TYPES:
        return PDF_POSITIVE_COLOR
    if forecast_type in NEGATIVE_TYPES:
        return PDF_NEGATIVE_COLOR
    return "#1F3864"


def _pdf_colored_bold(value, color):
    return f'<font color="{color}"><b>{_xml_escape(value)}</b></font>'


def _xml_escape(value):
    return (str(value).replace("&", "&amp;")
                      .replace("<", "&lt;")
                      .replace(">", "&gt;"))


def build_excel(date, items, out_path):
    items = [enrich_item_subset(dict(x)) for x in items]
    wb = Workbook()

    # Sheet 1: 汇总
    ws_sum = wb.active
    ws_sum.title = "汇总"
    ws_sum.append(["指标", "数值"])
    ws_sum.append(["披露家数", len(items)])
    ws_sum.append(["预增 / 续盈 / 扭亏",
                   sum(1 for x in items if x.get("预告类型") in ("预增", "续盈", "扭亏"))])
    ws_sum.append(["预减 / 首亏 / 续亏",
                   sum(1 for x in items if x.get("预告类型") in ("预减", "首亏", "续亏"))])
    ws_sum.append(["不确定",
                   sum(1 for x in items if x.get("预告类型") == "不确定")])
    ws_sum.append(["AI 主题相关",
                   sum(1 for x in items if x.get("AI主题相关") == 1)])
    for cell in ws_sum[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 16

    # Sheet 2: 明细
    cols = ["公告日期", "证券代码", "证券简称", "所属子集", "预告类型", "预告期间", "口径",
            "下限万元", "上限万元", "上年同期万元",
            "同比下限%", "同比上限%", "原因摘要", "AI主题相关", "AI 说明"]
    ws = wb.create_sheet("明细")
    # 按"同比上限%"降序排：增速最快在第一行
    def _upper(it):
        try:
            return float(it.get("同比上限%", 0) or 0)
        except (TypeError, ValueError):
            return 0.0
    items_sorted = sorted(items, key=_upper, reverse=True)
    ws.append(cols)
    for it in items_sorted:
        yoy_low = it.get("同比下限%", "")
        yoy_high = it.get("同比上限%", "")
        if isinstance(yoy_low, (int, float)):
            yoy_low = round(yoy_low, 2)
        if isinstance(yoy_high, (int, float)):
            yoy_high = round(yoy_high, 2)
        ws.append([
            it.get("公告日期", ""),
            it.get("证券代码", ""),
            it.get("证券简称", ""),
            it.get("所属子集", "其他"),
            it.get("预告类型", ""),
            it.get("预告期间", ""),
            it.get("口径", ""),
            it.get("下限万元", ""),
            it.get("上限万元", ""),
            it.get("上年同期万元", ""),
            yoy_low,
            yoy_high,
            it.get("原因摘要", ""),
            it.get("AI主题相关", 0),
            it.get("AI说明", ""),
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ai = row[13].value
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if ai == 1:
            for cell in row:
                cell.fill = AI_FILL
    widths = [12, 10, 14, 26, 10, 12, 12, 12, 12, 14, 10, 10, 40, 10, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)


def build_pdf(date, weekday, items, out_path, data_source="iFind 公告检索"):
    items = [enrich_item_subset(dict(x)) for x in items]
    doc = SimpleDocTemplate(
        str(out_path), pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=1.8 * cm, bottomMargin=1.8 * cm,
        title=f"A 股业绩预告日报 {date}",
    )
    base = getSampleStyleSheet()
    title_s = ParagraphStyle("T", parent=base["Heading1"],
                             fontName=FONT_BOLD, fontSize=18, spaceAfter=10,
                             textColor=colors.HexColor("#1F3864"), alignment=1)
    h2 = ParagraphStyle("H2", parent=base["Heading2"],
                        fontName=FONT_BOLD, fontSize=13, spaceAfter=4, spaceBefore=10,
                        textColor=colors.HexColor("#1F3864"))
    body = ParagraphStyle("B", parent=base["BodyText"], fontName=FONT_REG,
                          fontSize=10, leading=14, spaceAfter=2)
    center_body = ParagraphStyle("CB", parent=body, alignment=1)
    note = ParagraphStyle("N", parent=body, textColor=colors.grey, fontSize=6.5,
                          leading=9, alignment=1)

    flow = []
    flow.append(Paragraph(f"A 股业绩预告日报 · {date}（{weekday}）", title_s))
    flow.append(HRFlowable(width="100%", thickness=0.6, color=colors.black,
                           spaceBefore=0, spaceAfter=6))
    flow.append(Paragraph(f"1: 数据来源：{data_source}（关键词\"业绩预告\"），未经审计。", note))
    flow.append(Paragraph("2: 业绩口径：以下数据按公司公告披露原文摘录，单位：万元。", note))
    flow.append(Paragraph("3: 主题高亮：覆盖算力 / 芯片 / 光模块 / 大模型 / 应用 / 数据，相关公司加 [AI] 标记。", note))
    flow.append(Spacer(1, 0.4 * cm))

    # 汇总
    flow.append(Paragraph("一、当日汇总", h2))
    positive_label = f'<font color="{PDF_POSITIVE_COLOR}"><b>利好</b></font>'
    negative_label = f'<font color="{PDF_NEGATIVE_COLOR}"><b>利空</b></font>'
    header_label = '<font color="white"><b>{}</b></font>'
    positive_count = sum(1 for x in items if x.get("预告类型") in POSITIVE_TYPES)
    negative_count = sum(1 for x in items if x.get("预告类型") in NEGATIVE_TYPES)
    rows = [[Paragraph(header_label.format("指标"), center_body), Paragraph(header_label.format("数值"), center_body)],
            ["披露家数", str(len(items))],
            [Paragraph(positive_label, center_body), Paragraph(_pdf_colored_bold(positive_count, PDF_POSITIVE_COLOR), center_body)],
            [Paragraph(negative_label, center_body), Paragraph(_pdf_colored_bold(negative_count, PDF_NEGATIVE_COLOR), center_body)],
            ["不确定", str(sum(1 for x in items if x.get("预告类型") == "不确定"))],
            ["AI 主题相关", str(sum(1 for x in items if x.get("AI主题相关") == 1))]]
    t = Table(rows, colWidths=[6.5 * cm, 3.5 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), FONT_REG),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#F2F2F2"), colors.white]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    flow.append(t)
    flow.append(Spacer(1, 0.4 * cm))

    # 明细
    flow.append(Paragraph("二、明细", h2))
    if not items:
        flow.append(Paragraph("今日无业绩预告披露。", body))
    else:
        def append_detail_item(i, it):
            tag = "  [AI]" if it.get("AI主题相关") == 1 else ""
            name = _xml_escape(it.get("证券简称", ""))
            code = _xml_escape(it.get("证券代码", ""))
            name_color = _pdf_name_color(it.get("预告类型", ""))
            flow.append(Paragraph(
                f'{i}. <font color="{name_color}"><b>{name}（{code}）</b></font>{tag}', h2))
            last_year = it.get("上年同期万元", "")
            metric = it.get("口径", "归母净利润")
            range_label = "营业收入区间" if metric == "营业收入" else "净利润区间"
            profit_range = f"{it.get('下限万元', '')} ~ {it.get('上限万元', '')} 万元"
            if last_year not in ("", None):
                profit_range += f"（上年同期 {last_year} 万元）"
            yoy_low = it.get("同比下限%", "")
            yoy_high = it.get("同比上限%", "")
            if isinstance(yoy_low, (int, float)):
                yoy_low = round(yoy_low, 2)
            if isinstance(yoy_high, (int, float)):
                yoy_high = round(yoy_high, 2)
            yoy_text = it.get("同比变动说明") or f"{yoy_low}% ~ {yoy_high}%"
            detail = [
                ("公告日期", it.get("公告日期", "")),
                ("所属子集", it.get("所属子集", "其他")),
                ("预告类型", it.get("预告类型", "")),
                ("预告期间", it.get("预告期间", "")),
                ("口径", it.get("口径", "")),
                (range_label, profit_range),
                ("同比变动", yoy_text),
                ("业绩变动原因", it.get("原因摘要") or "公告未披露具体原因"),
            ]
            if it.get("AI主题相关") == 1:
                detail.append(("AI 大模型相关性", it.get("AI说明", "")))
            table_rows = []
            for k, v in detail:
                if k == "同比变动" and it.get("预告类型", "") in POSITIVE_TYPES | NEGATIVE_TYPES:
                    table_rows.append([
                        Paragraph(_pdf_colored_bold(k, name_color), body),
                        Paragraph(_pdf_colored_bold(v, name_color), body),
                    ])
                else:
                    table_rows.append([
                        Paragraph(f"<b>{_xml_escape(k)}</b>", body),
                        Paragraph(_xml_escape(v), body),
                    ])
            tbl = Table(
                table_rows,
                colWidths=[3.0 * cm, 13.0 * cm])
            tbl.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            flow.append(tbl)
            flow.append(Spacer(1, 0.2 * cm))

        positive_items = [x for x in items if x.get("预告类型") in POSITIVE_TYPES]
        negative_items = [x for x in items if x.get("预告类型") in NEGATIVE_TYPES]
        other_items = [x for x in items if x.get("预告类型") not in POSITIVE_TYPES | NEGATIVE_TYPES]

        flow.append(Paragraph("2.1 利好", h2))
        if positive_items:
            for i, it in enumerate(positive_items, 1):
                append_detail_item(i, it)
        else:
            flow.append(Paragraph("无。", body))

        flow.append(Paragraph("2.2 利空", h2))
        if negative_items:
            for i, it in enumerate(negative_items, 1):
                append_detail_item(i, it)
        else:
            flow.append(Paragraph("无。", body))

        if other_items:
            flow.append(Paragraph("2.3 其他", h2))
            for i, it in enumerate(other_items, 1):
                append_detail_item(i, it)

    doc.build(flow)


def main():
    if len(sys.argv) < 3:
        print("Usage: build_report.py <input.json> <output_dir>", file=sys.stderr)
        sys.exit(1)
    input_path = Path(sys.argv[1])
    output_dir = Path(sys.argv[2])
    output_dir.mkdir(parents=True, exist_ok=True)

    data = json.loads(input_path.read_text(encoding="utf-8"))
    date = data["date"]
    weekday = data.get("weekday", "")
    items = data.get("items", [])
    issues = []
    for item in items:
        issues.extend(check_item(item))
    if issues:
        print(f"[build_report] input quality check failed: {len(issues)} issues", file=sys.stderr)
        for issue in issues:
            print(f"- {issue}", file=sys.stderr)
        sys.exit(1)

    items = [enrich_item_subset(dict(x)) for x in items]
    data_source = data.get("data_source", "iFind 公告检索")

    xlsx = output_dir / f"A股业绩预告-{date}.xlsx"
    pdf = output_dir / f"A股业绩预告-{date}.pdf"
    build_excel(date, items, xlsx)
    build_pdf(date, weekday, items, pdf, data_source=data_source)
    print(f"OK: {xlsx}")
    print(f"OK: {pdf}")

if __name__ == "__main__":
    main()
